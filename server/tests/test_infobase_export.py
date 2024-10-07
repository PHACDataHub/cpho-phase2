from io import BytesIO

from django.urls import reverse

import openpyxl
from phac_aspc.rules import patch_rules

from cpho.model_factories import (
    BenchmarkingFactory,
    IndicatorDatumFactory,
    IndicatorFactory,
    TrendAnalysisFactory,
)
from cpho.models import (
    Benchmarking,
    Country,
    DimensionType,
    DimensionValue,
    Indicator,
    IndicatorDatum,
    Period,
    TrendAnalysis,
)
from cpho.services import (
    SubmitIndicatorDataService,
    SubmitIndicatorMetaDataService,
)


def test_infobase_export(vanilla_user_client):
    url = reverse("infobase_export")

    canada = Country.objects.get(name_en="Canada")

    indicators = IndicatorFactory.create_batch(5)
    for i in indicators:
        IndicatorDatumFactory.create_batch(5, indicator=i)

        Benchmarking.objects.create(
            indicator=i,
            oecd_country=canada,
            value=1,
            year=2020,
            methodology_differences="True",
        )
        TrendAnalysis.objects.create(
            indicator=i,
            data_point=1.1,
            year="2019-2020",
            trend=TrendAnalysis.TREND_CHOICES[1][0],
        )

    with patch_rules(is_admin_or_hso=False):
        response = vanilla_user_client.get(url)
        assert response.status_code == 403

    with patch_rules(is_admin_or_hso=True):
        response = vanilla_user_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/vnd.ms-excel"
        assert response.content


def test_infobase_export_queries(
    vanilla_user_client, django_assert_max_num_queries
):
    url = reverse("infobase_export")

    IndicatorFactory.create_batch(15)
    IndicatorDatumFactory.create_batch(15)
    TrendAnalysisFactory.create_batch(15)
    BenchmarkingFactory.create_batch(15)

    with patch_rules(is_admin_or_hso=True), django_assert_max_num_queries(10):
        response = vanilla_user_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/vnd.ms-excel"


def get_data_from_worksheet(file, sheet_num):
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[sheet_num]
    header_row = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(header_row, row)))
    return data


def test_infobase_export_with_approved_data(
    vanilla_user_client, vanilla_user, django_assert_max_num_queries
):
    url = reverse("infobase_export")
    sex_dim_type = DimensionType.objects.get(code="sex")
    possible_dimension_value = sex_dim_type.possible_values.all()
    p = Period.objects.first()

    i1 = IndicatorFactory.create(
        name="indicator1",
    )
    i2 = IndicatorFactory.create(
        name="indicator2",
    )
    male_i1 = IndicatorDatumFactory.create(
        indicator=i1,
        period=p,
        dimension_type=sex_dim_type,
        dimension_value=possible_dimension_value[0],
        value=1.1,
    )
    female_i1 = IndicatorDatumFactory.create(
        indicator=i1,
        period=p,
        dimension_type=sex_dim_type,
        dimension_value=possible_dimension_value[1],
        value=1.2,
    )
    male_i2 = IndicatorDatumFactory.create(
        indicator=i2,
        period=p,
        dimension_type=sex_dim_type,
        dimension_value=possible_dimension_value[0],
        value=2.1,
    )
    female_i2 = IndicatorDatumFactory.create(
        indicator=i2,
        period=p,
        dimension_type=sex_dim_type,
        dimension_value=possible_dimension_value[1],
        value=2.2,
    )
    i1 = Indicator.objects.get(pk=i1.pk)

    SubmitIndicatorDataService(
        i1, p, sex_dim_type, "hso", vanilla_user
    ).perform()

    # Indicator data submitted but Indicator metadata not submitted yet
    # Export should not include the data
    with patch_rules(is_admin_or_hso=True), django_assert_max_num_queries(15):
        response = vanilla_user_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/vnd.ms-excel"
        assert response.content
        file = BytesIO(response.content)
        data = get_data_from_worksheet(file, 0)
        assert len(data) == 0

    # indicator metadata now submitted
    SubmitIndicatorMetaDataService(i1, "hso", vanilla_user).perform()

    # sex data for i1 should be included in the export
    with patch_rules(is_admin_or_hso=True), django_assert_max_num_queries(15):
        response = vanilla_user_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/vnd.ms-excel"
        assert response.content
        file = BytesIO(response.content)

        # only i1 should be included in the indicator sheet
        data = get_data_from_worksheet(file, 0)
        assert len(data) == 1
        assert data[0]["name"] == "indicator1"

        # only i1 data should be included in the indicator_data sheet
        data = get_data_from_worksheet(file, 1)
        assert len(data) == 2
        for item in data:
            assert item["indicator name"] == "indicator1"

    # update name of i1
    i1.name = "indicator1_updated"
    i1.save()

    # indicator name update but not submitted
    with patch_rules(is_admin_or_hso=True), django_assert_max_num_queries(15):
        response = vanilla_user_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/vnd.ms-excel"
        assert response.content
        file = BytesIO(response.content)

        # only i1 should be included in the indicator sheet
        data = get_data_from_worksheet(file, 0)
        assert len(data) == 1
        # indicator name should not be updated
        assert data[0]["name"] == "indicator1"

        data = get_data_from_worksheet(file, 1)
        assert len(data) == 2
        for item in data:
            # indicator name should not be updated
            assert item["indicator name"] == "indicator1"
